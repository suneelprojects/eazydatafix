from openpyxl import Workbook
from openpyxl.styles import Font

from easydatafix.models.assessment_report import AssessmentReport


class ExcelReport:
    """
    Exports an AssessmentReport as an Excel report.
    """

    @staticmethod
    def export(
        report: AssessmentReport,
        output_file: str,
    ) -> None:

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Assessment Report"

        bold = Font(bold=True)

        row = 1

        # Dataset Information

        sheet.cell(row=row, column=1).value = "Dataset Information"
        sheet.cell(row=row, column=1).font = bold
        row += 1

        sheet.append(["Property", "Value"])

        sheet.append(
            [
                "File Name",
                report.dataset_info.file_name,
            ]
        )

        sheet.append(
            [
                "Rows",
                report.dataset_info.rows,
            ]
        )

        sheet.append(
            [
                "Columns",
                report.dataset_info.columns,
            ]
        )

        sheet.append(
            [
                "Memory Usage",
                f"{report.dataset_info.memory_usage_bytes:,} Bytes",
            ]
        )

        row = sheet.max_row + 2

        # Overall Quality

        sheet.cell(row=row, column=1).value = "Overall Quality"
        sheet.cell(row=row, column=1).font = bold
        row += 1

        sheet.append(
            [
                "Score",
                report.quality.score,
            ]
        )

        sheet.append(
            [
                "Grade",
                report.quality.grade,
            ]
        )

        row = sheet.max_row + 2

        # Quality Dimensions

        sheet.cell(row=row, column=1).value = "Quality Dimensions"
        sheet.cell(row=row, column=1).font = bold
        row += 1

        sheet.append(
            [
                "Metric",
                "Score",
            ]
        )

        sheet.append(
            [
                "Completeness",
                report.quality_dimensions.completeness,
            ]
        )

        sheet.append(
            [
                "Uniqueness",
                report.quality_dimensions.uniqueness,
            ]
        )

        sheet.append(
            [
                "Validity",
                report.quality_dimensions.validity,
            ]
        )

        sheet.append(
            [
                "Consistency",
                report.quality_dimensions.consistency,
            ]
        )

        sheet.append(
            [
                "Accuracy",
                report.quality_dimensions.accuracy,
            ]
        )

        sheet.append(
            [
                "Timeliness",
                report.quality_dimensions.timeliness,
            ]
        )

        sheet.append(
            [
                "Overall",
                report.quality_dimensions.overall,
            ]
        )

        row = sheet.max_row + 2

        # Recommendations

        sheet.cell(row=row, column=1).value = "Recommendations"
        sheet.cell(row=row, column=1).font = bold
        row += 1

        sheet.append(
            [
                "Priority",
                "Category",
                "Title",
                "Auto Fix",
                "Description",
            ]
        )

        for recommendation in report.recommendations:

            sheet.append(
                [
                    recommendation.priority,
                    recommendation.category,
                    recommendation.title,
                    "Yes"
                    if recommendation.auto_fix_available
                    else "No",
                    recommendation.description,
                ]
            )

        row = sheet.max_row + 2

        # Validations

        sheet.cell(row=row, column=1).value = "Validations"
        sheet.cell(row=row, column=1).font = bold
        row += 1

        sheet.append(
            [
                "Rule",
                "Column",
                "Status",
                "Message",
            ]
        )

        for validation in report.validations:

            sheet.append(
                [
                    validation.rule,
                    validation.column,
                    "PASS"
                    if validation.passed
                    else "FAIL",
                    validation.message,
                ]
            )

        # Auto-size columns

        for column_cells in sheet.columns:

            length = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column_cells
            )

            sheet.column_dimensions[
                column_cells[0].column_letter
            ].width = min(length + 3, 60)

        workbook.save(output_file)
